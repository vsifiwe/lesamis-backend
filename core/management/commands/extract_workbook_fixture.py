import hashlib
import json
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


FIXTURE_VERSION = 1
MEMBER_ALIASES = {
    'Chrisitan': 'Christian',
    'Obadias': 'Obadiah',
}
FRENCH_MONTHS = {
    'janvier': 1, 'fevrier': 2, 'février': 2, 'werurwe': 3, 'mars': 3,
    'avril': 4, 'mata': 4, 'mai': 5, 'juin': 6, 'kamena': 6,
    'juillet': 7, 'nyakanga': 7, 'aout': 8, 'août': 8, 'septembre': 9,
    'nzeri': 9, 'octobre': 10, 'novembre': 11, 'decembre': 12,
    'décembre': 12, 'ukuboza': 12, 'mutarama': 1,
}


def decimal_string(value):
    return str(Decimal(str(value or 0)).quantize(Decimal('0.01')))


def parse_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    match = re.match(r'(\d{1,2})\s+([^\s]+)\s+(\d{4})', text.lower())
    if match and match.group(2) in FRENCH_MONTHS:
        return date(int(match.group(3)), FRENCH_MONTHS[match.group(2)], int(match.group(1))).isoformat()
    raise CommandError(f'Cannot parse workbook date: {value!r}')


class WorkbookExtractor:
    def __init__(self, source):
        self.source = Path(source)
        self.values = load_workbook(self.source, data_only=True)
        self.formulas = load_workbook(self.source, data_only=False)

    def source_ref(self, sheet, cell, notes=''):
        value_cell = self.values[sheet][cell]
        formula_cell = self.formulas[sheet][cell]
        if isinstance(value_cell, tuple):
            values = [item.value for row in value_cell for item in row]
            formulas = [item.value for row in formula_cell for item in row]
            value = json.dumps(values, default=str, ensure_ascii=False)
            formula = json.dumps(formulas, default=str, ensure_ascii=False)
        else:
            value = value_cell.value
            formula = formula_cell.value
        return {
            'sheet': sheet,
            'cell': cell,
            'formula': formula if isinstance(formula, str) and formula.startswith('=') else '',
            'value': '' if value is None else str(value),
            'notes': notes,
        }

    def extract(self):
        summary = self.values["Vue d'Ensemble"]
        declared = datetime.strptime(summary['B3'].value.replace('Ku wa ', ''), '%d/%m/%Y').date()
        members, contributions = self.extract_contributions()
        loans = self.extract_loans()
        social_expenses = self.extract_social_expenses()
        return {
            'fixture_version': FIXTURE_VERSION,
            'source_name': self.source.name,
            'source_sha256': hashlib.sha256(self.source.read_bytes()).hexdigest(),
            'declared_summary_date': declared.isoformat(),
            'latest_transaction_date': max(
                parse_date(row['date']) for row in loans['active'] if row['date']
            ),
            'members': members,
            'contributions': contributions,
            'loans': loans,
            'social_expenses': social_expenses,
            'income': self.extract_income(),
            'bank_charges': [{
                'amount': decimal_string(summary['C19'].value),
                'description': summary['B19'].value,
                'date': None,
                'date_precision': 'unknown',
                'source': self.source_ref("Vue d'Ensemble", 'C19'),
            }],
            'investments': self.extract_investments(),
            'reconciliation': {
                'as_of_date': declared.isoformat(),
                'calculated_cash_balance': decimal_string(summary['C39'].value),
                'stated_bank_balance': decimal_string(summary['E39'].value),
                'variance': decimal_string(summary['G39'].value),
                'expected_total_assets': decimal_string(summary['C40'].value),
                'source': self.source_ref("Vue d'Ensemble", 'C39', 'Calculated cash; E39 is stated bank balance and G39 is variance.'),
            },
            'targets': {
                'contribution_capital': decimal_string(summary['C15'].value),
                'contribution_social': decimal_string(summary['C6'].value),
                'contribution_social_plus': decimal_string(summary['C9'].value),
                'social_expenses': decimal_string(summary['C7'].value),
                'social_plus_expenses': decimal_string(summary['C10'].value),
                'bank_interest': decimal_string(summary['C16'].value),
                'joining_fees': decimal_string(summary['C17'].value),
                'collected_penalties': decimal_string(summary['C18'].value),
                'bank_charges': decimal_string(summary['C19'].value),
                'loan_principal': decimal_string(summary['C22'].value),
                'loan_interest': decimal_string(summary['C23'].value),
                'loan_repayments': decimal_string(summary['C24'].value),
                'active_loan_principal': decimal_string(self.values['Details de credits']['G52'].value),
                'active_loan_outstanding': decimal_string(self.values['Details de credits']['L52'].value),
                'investments': decimal_string(sum(Decimal(x['amount']) for x in self.extract_investments())),
                'investment_profits': decimal_string(sum(Decimal(p['amount']) for x in self.extract_investments() for p in x['profits'])),
                'calculated_cash_balance': decimal_string(summary['C39'].value),
                'expected_total_assets': decimal_string(summary['C40'].value),
                'shares': int(self.values['Details']['V53'].value),
            },
        }

    def extract_contributions(self):
        sheet = self.values['Contribution Individuel Mensuel']
        formula_sheet = self.formulas['Contribution Individuel Mensuel']
        periods = []
        year = 2020
        month = 8
        while (year, month) <= (2021, 4):
            periods.append((year, month, ('capital', 'social')))
            year, month = (year + 1, 1) if month == 12 else (year, month + 1)
        year, month = 2021, 5
        while (year, month) <= (2026, 12):
            periods.append((year, month, ('capital', 'social', 'social_plus')))
            year, month = (year + 1, 1) if month == 12 else (year, month + 1)

        details = self.values['Details']
        members = []
        contributions = []
        column = 3
        for row in range(3, 52):
            source_name = sheet.cell(row, 2).value
            canonical = MEMBER_ALIASES.get(source_name, source_name)
            first_period = None
            member_number = f'MBR-{row - 2:04d}'
            row_contributions = []
            scan_column = 3
            for period_year, period_month, funds in periods:
                for offset, fund in enumerate(funds):
                    cell = sheet.cell(row, scan_column + offset)
                    if cell.value not in (None, 0):
                        first_period = first_period or date(period_year, period_month, 1)
                        formula_cell = formula_sheet[cell.coordinate]
                        row_contributions.append({
                            'member_number': member_number,
                            'year': period_year,
                            'month': period_month,
                            'fund': fund,
                            'amount': decimal_string(cell.value),
                            'source': {
                                'sheet': sheet.title,
                                'cell': cell.coordinate,
                                'formula': formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith('=') else '',
                                'value': str(cell.value),
                                'notes': '',
                            },
                        })
                scan_column += len(funds)
            contributions.extend(row_contributions)
            members.append({
                'member_number': member_number,
                'source_name': source_name,
                'first_name': canonical,
                'last_name': '',
                'aliases': sorted(set([source_name]) - {canonical}),
                'join_date': (first_period or date(2026, 1, 1)).isoformat(),
                'share_count': int(details.cell(row + 1, 22).value or 0),
                'source': self.source_ref('Details', f'V{row + 1}'),
            })
        return members, contributions

    def extract_loans(self):
        sheet = self.values['Details de credits']
        active = []
        historical = []
        for row in range(3, 52):
            source_name = sheet.cell(row, 3).value
            if not source_name:
                continue
            canonical = MEMBER_ALIASES.get(source_name, source_name)
            total_principal = Decimal(str(sheet.cell(row, 4).value or 0))
            total_interest = Decimal(str(sheet.cell(row, 5).value or 0))
            total_paid = Decimal(str(sheet.cell(row, 6).value or 0))
            active_principal = Decimal(str(sheet.cell(row, 7).value or 0))
            duration = int(sheet.cell(row, 8).value or 0)
            rate_fraction = Decimal(str(sheet.cell(row, 9).value or 0))
            active_interest = active_principal * rate_fraction
            active_paid = Decimal(str(sheet.cell(row, 11).value or 0))
            active_outstanding = Decimal(str(sheet.cell(row, 12).value or 0))
            if active_principal:
                active.append({
                    'member_name': canonical,
                    'principal': decimal_string(active_principal),
                    'interest': decimal_string(active_interest),
                    'paid': decimal_string(active_paid),
                    'outstanding': decimal_string(active_outstanding),
                    'duration_months': duration,
                    'rate_percent': decimal_string(rate_fraction * 100),
                    'date': parse_date(sheet.cell(row, 13).value),
                    'maturity_date': parse_date(sheet.cell(row, 14).value),
                    'source': self.source_ref('Details de credits', f'G{row}:N{row}'),
                })
            opening_principal = total_principal - active_principal
            opening_interest = total_interest - active_interest
            opening_paid = total_paid - active_paid
            if opening_principal or opening_interest or opening_paid:
                historical.append({
                    'member_name': canonical,
                    'principal': decimal_string(opening_principal),
                    'interest': decimal_string(opening_interest),
                    'paid': decimal_string(opening_paid),
                    'source': self.source_ref('Details de credits', f'D{row}:F{row}'),
                })
        return {'active': active, 'historical': historical}

    def extract_social_expenses(self):
        rows = []
        details_sources = [
            ('SOCIAL', 'E57', 'Ayakoreshejwe muri social'), ('SOCIAL', 'G57', 'Guhemba Sandrine'),
            ('SOCIAL', 'H57', 'Muhazi'), ('SOCIAL', 'I57', 'Guhemba Diane & Thadee'),
            ('SOCIAL', 'G59', 'Gufata mu mugongo Clement'), ('SOCIAL', 'I59', 'Guhemba Tuyizere & Edissa'),
            ('SOCIAL', 'K57', 'Guhemba Venuste'), ('SOCIAL', 'K59', 'Gutwerera Semanzi'),
            ('SOCIAL', 'K61', 'Guhemba Yvonne'), ('SOCIAL', 'U57', 'Guhemba Arnauld & Joelle'),
            ('SOCIAL', 'U63', 'Intwererano Vital'), ('SOCIAL', 'U65', 'Guhemba Phoebe & Semanzi'),
            ('SOCIAL', 'S57', 'Gutwerera Vanessa Mupenzi'), ('SOCIAL', 'L57', 'Guhemba Tuyizere & Edissa Mars 2024'),
            ('SOCIAL', 'L59', 'Cadeau Tuyizere & Edissa Mars 2024'),
            ('SOCIAL_PLUS', 'U59', 'Inama rusange du 24 Juin'), ('SOCIAL_PLUS', 'U61', 'Inama rusange du 24 Juin'),
            ('SOCIAL_PLUS', 'U67', 'Inama rusange du 22 Decembre'), ('SOCIAL_PLUS', 'U68', 'Inama rusange du 22 Decembre supplement'),
            ('SOCIAL_PLUS', 'U70', 'Avance sortie 27 Janvier 2024'), ('SOCIAL_PLUS', 'U72', 'Inama 11 Gicurasi 2024'),
        ]
        for fund, cell, name in details_sources:
            value = self.values['Details'][cell].value
            rows.append({'fund': fund, 'date': None, 'date_precision': 'unknown', 'category': 'support' if fund == 'SOCIAL' else 'event',
                         'name': name, 'amount': decimal_string(value), 'source': self.source_ref('Details', cell)})
        rows.append({'fund': 'SOCIAL_PLUS', 'date': None, 'date_precision': 'unknown', 'category': 'event',
                     'name': 'Kwidagadura hardcoded expenses', 'amount': decimal_string(1948000),
                     'source': self.source_ref("Vue d'Ensemble", 'C10', 'Hardcoded formula constants total 1,948,000.')})
        social = self.values['Social Loisirs']
        for row in range(3, 23):
            for fund, date_col, name_col, amount_col in [('SOCIAL', 1, 2, 3), ('SOCIAL_PLUS', 5, 6, 7)]:
                amount = social.cell(row, amount_col).value
                if amount:
                    rows.append({
                        'fund': fund, 'date': parse_date(social.cell(row, date_col).value), 'date_precision': 'exact',
                        'category': 'support' if fund == 'SOCIAL' else 'event',
                        'name': social.cell(row, name_col).value, 'amount': decimal_string(amount),
                        'source': self.source_ref('Social Loisirs', f'{get_column_letter(amount_col)}{row}'),
                    })
        return rows

    def extract_income(self):
        details = self.values['Details']
        summary = self.values["Vue d'Ensemble"]
        return [
            {'type': 'bank_interest', 'amount': decimal_string(summary['C16'].value), 'source': self.source_ref("Vue d'Ensemble", 'C16')},
            {'type': 'joining_fee', 'amount': decimal_string(summary['C17'].value), 'source': self.source_ref("Vue d'Ensemble", 'C17')},
            {'type': 'contribution_penalty', 'amount': decimal_string(details['H53'].value), 'source': self.source_ref('Details', 'H53')},
            {'type': 'loan_exit_penalty', 'amount': decimal_string(details['J53'].value), 'source': self.source_ref('Details', 'J53')},
        ]

    def extract_investments(self):
        return [
            {'name': 'BRD Sustainability Bond August 2024', 'type': 'bond', 'date': '2024-08-01', 'amount': decimal_string(20000000),
             'rate': '12.90', 'vesting_months': 84, 'source': self.source_ref("Vue d'Ensemble", 'C27'),
             'profits': [{'date': d, 'amount': decimal_string(1225500), 'source': self.source_ref("Vue d'Ensemble", 'C28')} for d in ('2025-03-31', '2025-09-30', '2026-03-31')]},
            {'name': 'Treasury Bond April 2025', 'type': 'bond', 'date': '2025-04-01', 'amount': decimal_string(10000000),
             'rate': '12.15', 'vesting_months': 120, 'source': self.source_ref("Vue d'Ensemble", 'C29'),
             'profits': [{'date': d, 'amount': decimal_string(607500), 'source': self.source_ref("Vue d'Ensemble", 'C30')} for d in ('2025-10-31', '2026-04-30')]},
            {'name': 'Treasury Bond May 2025', 'type': 'bond', 'date': '2025-05-01', 'amount': decimal_string(self.values["Vue d'Ensemble"]['C31'].value),
             'rate': '13.27', 'vesting_months': 240, 'source': self.source_ref("Vue d'Ensemble", 'C31'),
             'profits': [{'date': d, 'amount': decimal_string(729850), 'source': self.source_ref("Vue d'Ensemble", 'C32')} for d in ('2025-06-30', '2025-12-31')]},
            {'name': 'Treasury Bond Jan 2026', 'type': 'bond', 'date': '2026-01-01', 'amount': decimal_string(10000000),
             'rate': '12.15', 'vesting_months': 120, 'source': self.source_ref("Vue d'Ensemble", 'C33'), 'profits': []},
        ]


class Command(BaseCommand):
    help = 'Extract the authoritative workbook into a deterministic normalized JSON fixture.'

    def add_arguments(self, parser):
        parser.add_argument('workbook')
        parser.add_argument('--output', default='core/fixtures/workbook_import_v1.json')

    def handle(self, *args, **options):
        extractor = WorkbookExtractor(options['workbook'])
        data = extractor.extract()
        output = Path(options['output'])
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(json.dumps(data, indent=2, sort_keys=True, ensure_ascii=False) + '\n', encoding='utf-8')
        self.stdout.write(self.style.SUCCESS(f'Wrote normalized fixture to {output}'))
